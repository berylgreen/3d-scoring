import os
import sys
import json
import pandas as pd
from pathlib import Path

# 引入项目配置
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from core.config import settings
from core.logger import logger

def main():
    json_path = settings.GRADING_RESULTS_JSON
    
    if not json_path.exists():
        logger.error(f"错误: 找不到文件 {json_path}")
        return
    
    logger.info(f"正在读取 JSON 数据: {json_path}")
    with open(json_path, 'r', encoding='utf-8') as f:
        all_results = json.load(f)
        
    if not all_results:
        logger.warning("JSON 文件中没有数据！")
        return
        
    # 解析嵌套的 JSON 数据并展平
    flattened_data = []
    for item in all_results:
        group_info = item.get('group_info', {})
        individuals = item.get('individuals', [])
        scores = group_info.get('scores', {})
        
        group_base = {
            '小组名称': group_info.get('group_name', ''),
            '小组': group_info.get('total_group_score', 0),
        }
        
        # 兼容多种评分标准（建模/动画）
        score_mapping = {
            # 建模类
            'theme_creativity': '主题创意',
            'difficulty_workload': '难度与工作量',
            'modeling_accuracy': '建模准确性与完成度',
            'model_details': '模型细节与表现力',
            'topology': '拓扑结构合理性',
            'materials_textures': '材质和贴图',
            'uv_mapping': '纹理映射/UV展开',
            'lighting_rendering': '光照渲染',
            'visual_quality': '整体视觉质量',
            'documentation': '文档整理',
            
            # 动画类
            'creativity': '创意与剧本',
            'storyboard': '分镜与镜头',
            'modeling': '建模与场景',
            'basic_tech': '基础动画技术',
            'adv_tech': '高级动画技术',
            'fluency': '动画流畅度',
            'rendering': '材质光影与渲染',
            'post_production': '剪辑特效与音效配乐',
            'document': '文档整理'
        }
        
        for key, value in scores.items():
            if key in score_mapping:
                chinese_name = score_mapping[key]
                group_base[chinese_name] = value

        group_base.update({
            '小组工作量评价': group_info.get('workload_comment', ''),
            '小组整体评语': group_info.get('comments', ''),
            '作品路径': group_info.get('folder_path', '')
        })
        
        if not individuals:
            # 如果没有解析出个人，仅保留组信息
            row = {
                '学号': '',
                '姓名': '',
                '期末成绩(100)': group_base['小组'] * 0.8,
                '平时成绩(100)': '',
                '总分': '',
                '个人': 0,
                '个人任务描述': '',
                '个人评语': ''
            }
            row.update(group_base)
            flattened_data.append(row)
        else:
            for ind in individuals:
                ind_score = ind.get('individual_score', 0)
                group_score = group_base['小组']
                total_score = group_score * 0.8 + ind_score
                
                row = {
                    '学号': ind.get('student_id', ''),
                    '姓名': ind.get('student_name', ''),
                    '期末成绩(100)': total_score,
                    '平时成绩(100)': '',
                    '总分': '',
                    '个人': ind_score,
                }
                row.update(group_base)
                row['个人任务描述'] = ind.get('task_description', '')
                row['个人评语'] = ind.get('individual_comment', '')
                
                flattened_data.append(row)
                
    df = pd.DataFrame(flattened_data)
    
    sub_items = [
        # 建模
        '主题创意', '难度与工作量', '建模准确性与完成度', '模型细节与表现力',
        '拓扑结构合理性', '材质和贴图', '纹理映射/UV展开', '光照渲染',
        
        # 动画
        '创意与剧本', '分镜与镜头', '建模与场景', '基础动画技术',
        '高级动画技术', '动画流畅度', '材质光影与渲染', '剪辑特效与音效配乐',
        
        # 通用
        '整体视觉质量', '文档整理'
    ]
    
    # 调整列的顺序，使其更加符合教师查看成绩的习惯
    cols = ['学号', '姓名', '小组名称'] + sub_items + [
            '小组', '个人', '期末成绩(100)', '平时成绩(100)', '总分',
            '个人任务描述', '个人评语', 
            '小组工作量评价', '小组整体评语', '作品路径']
    
    existing_cols = [c for c in cols if c in df.columns]
    df = df[existing_cols]
    
    # === 自动寻找名单表格进行排序 ===
    data_dir = settings.DATA_DIR
    roster_file = None
    for f in data_dir.glob("*.xlsx"):
        if not f.name.startswith("~$"):
            roster_file = f
            break
            
    if roster_file:
        logger.info(f"读取到学生名单: {roster_file.name}，正在按名单顺序排序...")
        try:
            roster_df = pd.read_excel(roster_file)
            
            # 找到包含“学号”和“姓名”的列
            student_id_col = next((col for col in roster_df.columns if '学号' in str(col)), None)
            name_col = next((col for col in roster_df.columns if '姓名' in str(col)), None)
            
            if student_id_col:
                roster_df[student_id_col] = roster_df[student_id_col].astype(str).str.strip()
                df['学号'] = df['学号'].astype(str).str.strip()
                
                roster_df['__order__'] = range(len(roster_df))
                
                # 为了防止 merge 时的列名冲突（比如两边都有'学号'、'姓名'），我们先将 df 中的改名
                df_temp = df.copy()
                df_temp.rename(columns={'学号': '__json_id__', '姓名': '__json_name__'}, inplace=True)
                
                # 要提取的名单列
                roster_extract_cols = [student_id_col, '__order__']
                if name_col and name_col != student_id_col:
                    roster_extract_cols.append(name_col)
                    
                # 左连接名单
                df_merged = pd.merge(roster_df[roster_extract_cols], 
                                     df_temp, left_on=student_id_col, right_on='__json_id__', how='left')
                
                # 找出在评分数据里，但不在名单里的学生 (例如学号写错的)
                missing_in_roster = df_temp[~df_temp['__json_id__'].isin(roster_df[student_id_col])].copy()
                if not missing_in_roster.empty:
                    missing_in_roster['__order__'] = 999999
                    # 将 json 里的学号和姓名赋给对应的 roster 列，以便统一
                    missing_in_roster[student_id_col] = missing_in_roster['__json_id__']
                    if name_col:
                        missing_in_roster[name_col] = missing_in_roster['__json_name__']
                    df_merged = pd.concat([df_merged, missing_in_roster], ignore_index=True)
                
                # 排序
                df_merged = df_merged.sort_values(by='__order__')
                
                # 恢复标准列名，以名单的列优先作为输出的“学号”和“姓名”
                df_merged.rename(columns={student_id_col: '学号'}, inplace=True)
                if name_col:
                    df_merged.rename(columns={name_col: '姓名'}, inplace=True)
                else:
                    df_merged['姓名'] = df_merged['__json_name__']
                
                # 清理临时列，恢复原始的 cols 顺序
                df = df_merged[existing_cols]
                logger.info("已成功按学生名单顺序排版，并修复了姓名显示问题。")
            else:
                logger.warning(f"名单文件 {roster_file.name} 中未找到'学号'列，无法执行排序。")
        except Exception as e:
            logger.error(f"按名单排序时出错: {e}")
            
    # 导出到 Excel
    excel_path = settings.EXCEL_PATH
    if not excel_path:
        excel_path = settings.RESULT_DIR / "评分记录表.xlsx"
        
    logger.info(f"正在生成精细化 Excel 表格及成绩统计...")
    
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # 1. 写入明细表
            df.to_excel(writer, sheet_name='学生成绩明细', index=False)
            
            # 2. 写入统计分析表
            if not df.empty:
                score_cols = [c for c in existing_cols if '分' in c or c in sub_items]
                if score_cols:
                    # 将需要统计的列转换为数值类型
                    for col in score_cols:
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                        
                    stats_df = df[score_cols].agg(['mean', 'max', 'min']).round(2)
                    stats_df.index = ['平均分', '最高分', '最低分']
                    stats_df = stats_df.T
                    stats_df.index.name = '评分项目'
                    stats_df.reset_index(inplace=True)
                    stats_df.to_excel(writer, sheet_name='成绩统计分析', index=False)
                    
                # 分数段统计
                if '期末成绩(100)' in df.columns:
                    total_scores = df['期末成绩(100)']
                    bins = [-1, 59.99, 69.99, 79.99, 89.99, 100]
                    labels = ['不及格(<60)', '及格(60-69)', '中等(70-79)', '良好(80-89)', '优秀(90-100)']
                    dist = pd.cut(total_scores, bins=bins, labels=labels).value_counts().sort_index()
                    dist_df = pd.DataFrame({'分数段': dist.index, '人数': dist.values})
                    dist_df.to_excel(writer, sheet_name='成绩统计分析', startrow=len(stats_df) + 3 if score_cols else 0, index=False)
            
            # 3. 格式美化：调整列宽
            try:
                worksheet = writer.sheets['学生成绩明细']
                for col_idx, col_name in enumerate(existing_cols, 1):
                    # 获取列字母
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col_idx)
                    
                    if '评语' in col_name or '描述' in col_name or '路径' in col_name:
                        worksheet.column_dimensions[col_letter].width = 40
                    elif col_name in sub_items:
                        worksheet.column_dimensions[col_letter].width = 4
                    elif '得分' in col_name:
                        worksheet.column_dimensions[col_letter].width = 8
                    elif '名称' in col_name:
                        worksheet.column_dimensions[col_letter].width = 25
                    else:
                        worksheet.column_dimensions[col_letter].width = 12
                        
                # 取消自动换行，保持表头只有一行高
                from openpyxl.styles import Alignment
                for cell in worksheet[1]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                worksheet.row_dimensions[1].height = 20
                
                # 写入期末成绩的 Excel 公式
                try:
                    if '期末成绩(100)' in existing_cols and '小组' in existing_cols and '个人' in existing_cols:
                        final_col = get_column_letter(existing_cols.index('期末成绩(100)') + 1)
                        group_col = get_column_letter(existing_cols.index('小组') + 1)
                        ind_col = get_column_letter(existing_cols.index('个人') + 1)
                        
                        has_pingshi = '平时成绩(100)' in existing_cols
                        has_total = '总分' in existing_cols
                        if has_pingshi and has_total:
                            pingshi_col = get_column_letter(existing_cols.index('平时成绩(100)') + 1)
                            total_col = get_column_letter(existing_cols.index('总分') + 1)

                        for row in range(2, len(df) + 2):
                            worksheet[f"{final_col}{row}"].value = f"=ROUND({group_col}{row}*0.8+{ind_col}{row}, 0)"
                            if has_pingshi and has_total:
                                # 默认总分为期末70%+平时30%，四舍五入。
                                worksheet[f"{total_col}{row}"].value = f"=IF(ISBLANK({pingshi_col}{row}), {final_col}{row}, ROUND({final_col}{row}*0.7+{pingshi_col}{row}*0.3, 0))"
                except Exception as e:
                    logger.error(f"写入公式失败: {e}")
                    
            except ImportError:
                pass # 如果 openpyxl.utils 导入失败，忽略列宽调整
                
        logger.info(f"成功！Excel 表格已保存至: {excel_path}")
        
    except Exception as e:
        logger.error(f"生成 Excel 失败: {e}")
        # 如果 openpyxl 等引擎报错，降级使用基本导出
        logger.info("尝试使用基础模式导出...")
        df.to_excel(excel_path, index=False)
        logger.info(f"基础 Excel 表格已保存至: {excel_path}")

if __name__ == "__main__":
    main()
